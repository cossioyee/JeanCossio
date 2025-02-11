def application(environ, start_response):
    start_response('200 OK', [('Content-Type', 'text/plain')])
    return [b'Hola mundo!']

import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# Ruta del archivo original y la nueva ruta
original_file_path = 'C:/Users/jeacossio/Downloads/ManagedSystems.xlsx'
new_file_path = 'C:/Users/jeacossio/OneDrive - bgeneral.com/Documentos/Nuevo Backup/Liberaciones/ManagedSystems.xlsx'

# Copiar el archivo a la nueva ubicación
shutil.copyfile(original_file_path, new_file_path)

# Eliminar el archivo original
if os.path.exists(original_file_path):
    os.remove(original_file_path)

# Cargar el archivo Excel
sheet_name = 'ManageEngine Patch Manager Plus'
df = pd.read_excel(new_file_path, sheet_name=sheet_name)

# Filtrar las filas que contienen "Stella" en la columna "Area Responsable"
df_filtered = df[df['Area Responsable'].str.contains('Stella', na=False)].copy()

# Convertir la columna "Computer Name" a minúsculas
df_filtered.loc[:, 'Computer Name'] = df_filtered['Computer Name'].str.lower()

# Clasificar la columna "Computer Name"
def classify_computer_name(name):
    if name.startswith('bgcd'):
        return 'dev'
    elif name.startswith('bgcp'):
        return 'amb'
    elif name.startswith('bgca'):
        return 'dr'
    elif name.startswith('bgdc'):
        return 'prod'
    else:
        return 'unknown'

df_filtered.loc[:, 'Classification'] = df_filtered['Computer Name'].apply(classify_computer_name)

# Clasificar la columna "Computer Name" por tecnología
def classify_tech(name):
    if 'ntps' in name:
        return 'ntp'
    if 'dynatrace' in name:
        return 'dynatrace'
    if 'nortexdc' in name:
        return 'decision center'
    if 'dsg' in name:
        return 'designer'
    if 'log' in name:
        return 'ELK'
    if 'lic' in name:
        return 'designer licencia'
    if 'hpv' in name or 'hyper' in name:
        return 'hypervisor'
    if 'rh0' in name:
        return 'hosts'
    if 'mon' in name:
        return 'monitoreo'
    if 'odmrs' in name:
        return 'odm rs'
    if 'odmbd' in name:
        return 'odm db'
    if 'xace' in name:
        return 'ace'
    if 'ecrionap' in name:
        return 'ecrion ap'
    if 'ecriondb' in name or 'ecrionbd' in name:
        return 'ecrion db'
    if 'engagelb' in name:
        return 'engage lb'
    if 'engagedb' in name or 'engdb' in name:
        return 'engage db'
    if 'engageap' in name or 'engap' in name or 'engagecap' in name:
        return 'engage ap'
    if 'lb' in name or 'ngx' in name:
        return 'nginx'
    if 'cg' in name or 'cgen' in name:
        return 'jboss / tomcat'
    if 'xels' in name or 'xes' in name or 'aes' in name:
        return 'elasticsearch'
    if 'camdb' in name or 'db' in name:
        return 'db'
    if 'api' in name or 'apr' in name:
        return 'container'
    if 'ap' in name:
        return 'ap liferay'
    else:
        return 'unknown'

df_filtered.loc[:, 'Tech'] = df_filtered['Computer Name'].apply(classify_tech)

# Convertir la columna "Last Patched Time" a formato fecha (día, mes y año con el mes en palabras)
df_filtered['Last Patched Time'] = pd.to_datetime(df_filtered['Last Patched Time'], errors='coerce').dt.strftime('%Y-%m-%d')
df_filtered['Last Boot Time'] = pd.to_datetime(df_filtered['Last Boot Time'], errors='coerce').dt.strftime('%Y-%m-%d')
df_filtered['Last Patched Month'] = pd.to_datetime(df_filtered['Last Patched Time'], errors='coerce').dt.strftime('%B')
df_filtered['Last Boot Month'] = pd.to_datetime(df_filtered['Last Boot Time'], errors='coerce').dt.strftime('%B')

# Eliminar las columnas especificadas
columns_to_drop = ['Area Responsable', 'Last Successful Scan', 'Remote Office', 'Area Responsable AP']
df_filtered.drop(columns=columns_to_drop, inplace=True)

# Guardar el DataFrame filtrado en una nueva hoja llamada "Stella"
with pd.ExcelWriter(new_file_path, mode='a', engine='openpyxl') as writer:
    df_filtered.to_excel(writer, sheet_name='Stella', index=False)

# Cargar el archivo Excel con openpyxl
wb = load_workbook(new_file_path)
ws = wb['Stella']

# Obtener el mes actual
current_month = pd.Timestamp.now().strftime('%B')
previous_month = (pd.Timestamp.now() - pd.DateOffset(months=1)).strftime('%B')

# Definir los colores
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Obtener el índice de la columna "Last Patched Month"
lpm_col_index = None
for idx, cell in enumerate(ws[1]):
    if cell.value == 'Last Patched Month':
        lpm_col_index = idx + 1
        break

# Aplicar el color a la columna "Last Patched Month"
for row in ws.iter_rows(min_row=2, min_col=lpm_col_index, max_col=lpm_col_index):
    for cell in row:
        if cell.value == current_month:
            cell.fill = green_fill
        elif cell.value == previous_month:
            cell.fill = yellow_fill
        else:
            cell.fill = red_fill

# Aplicar el color a la columna "Last Boot Month"
for row in ws.iter_rows(min_row=2, min_col=ws.max_column, max_col=ws.max_column):
    for cell in row:
        if cell.value == current_month:
            cell.fill = green_fill
        elif cell.value == previous_month:
            cell.fill = yellow_fill
        else:
            cell.fill = red_fill

# Obtener el índice de la columna "Health"
health_col_index = None
for idx, cell in enumerate(ws[1]):
    if cell.value == 'Health':
        health_col_index = idx + 1
        break

# Aplicar el color a la columna "Health"
if health_col_index:
    for row in ws.iter_rows(min_row=2, min_col=health_col_index, max_col=health_col_index):
        for cell in row:
            if cell.value == 'Healthy':
                cell.fill = green_fill
            elif cell.value == 'Vulnerable':
                cell.fill = yellow_fill
            elif cell.value == 'Highly Vulnerable':
                cell.fill = red_fill

# Guardar los cambios en el archivo Excel
wb.save(new_file_path)
