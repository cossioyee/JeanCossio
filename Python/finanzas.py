import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
from openpyxl.formatting.rule import ColorScaleRule

# Ruta del archivo original y la nueva ruta
# original_file_path = 'C:/Users/jeacossio/Downloads/ManagedSystems.xlsx'
# new_file_path = 'C:/Users/jeacossio/OneDrive - bgeneral.com/Documentos/Nuevo Backup/Liberaciones/ManagedSystems.xlsx'
old_file_path = 'D:/2021/importante-viejo.xlsx'
original_file_path = 'D:/2021/importante-nuevo.xlsx'
new_file_path = 'D:/2021/estado-super-general.xlsx'

# Eliminar el archivo original
if os.path.exists(new_file_path):
    os.remove(new_file_path)

# Copiar el archivo a la nueva ubicación
shutil.copyfile(original_file_path, new_file_path)

# Cargar el archivo Excel viejo
old_df_h1 = pd.read_excel(old_file_path, sheet_name='Hoja1')
old_df_h2 = pd.read_excel(old_file_path, sheet_name='Hoja2')

# Cargar el archivo Excel
sheet_name = 'Hoja1'
df = pd.read_excel(new_file_path, sheet_name=sheet_name)

# Agregar todo el old_df después del df
df = pd.concat([df, old_df_h2], ignore_index=True)
df = pd.concat([df, old_df_h1], ignore_index=True)

# Convertir la columna "Fecha" a formato fecha (día, mes y año con el mes en palabras)
df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce').dt.strftime('%d %B %Y')

# # Convertir las columnas "Monto" y "Saldo total" a formato moneda
# df['Monto'] = df['Monto'].apply(lambda x: "${:,.2f}".format(x))
# df['Saldo total'] = df['Saldo total'].apply(lambda x: "${:,.2f}".format(x))

# Eliminar la columna "Referencia"
if 'Referencia' in df.columns:
    df.drop(columns=['Referencia'], inplace=True)

# Separar la columna "Monto" en "Saliente" y "Entrante"
df['Monto'] = df['Monto'].replace('[\$,]', '', regex=True).astype(float)
df['Saliente'] = df['Monto'].apply(lambda x: x if x < 0 else None)
df['Entrante'] = df['Monto'].apply(lambda x: x if x > 0 else None)
df['Saliente'].fillna(0, inplace=True)
df['Entrante'].fillna(0, inplace=True)
# df['Monto'] = df['Monto'].apply(lambda x: "${:,.2f}".format(x))
# df['Saliente'] = df['Saliente'].apply(lambda x: "${:,.2f}".format(x))
# df['Entrante'] = df['Entrante'].apply(lambda x: "${:,.2f}".format(x))

# Insertar 4 filas y desplazar hacia abajo en cada cambio de mes en la columna "Fecha"
df['Mes'] = pd.to_datetime(df['Fecha'], format='%d %B %Y').dt.month
df['Año'] = pd.to_datetime(df['Fecha'], format='%d %B %Y').dt.year


# Eliminar la columna "Monto"
df.drop(columns=['Monto'], inplace=True)

# Mover la columna "Saldo total" al final
saldo_total = df.pop('Saldo total')
df['Saldo total'] = saldo_total

# Insertar filas en cada cambio de mes y sumar Saliente y Entrante
new_rows = []
previous_month = None
previous_year = None

for index, row in df.iterrows():
    current_month = row['Mes']
    current_year = row['Año']
    if previous_month is not None and (current_month != previous_month or current_year != previous_year):
        for _ in range(4):
            new_rows.append(pd.Series([None] * len(df.columns), index=df.columns))
    new_rows.append(row)
    previous_month = current_month
    previous_year = current_year

df = pd.DataFrame(new_rows)

# Eliminar las columnas temporales
df.drop(columns=['Mes', 'Año'], inplace=True)

# Guardar el DataFrame actualizado en la hoja "Estado"
with pd.ExcelWriter(new_file_path, mode='a', engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Estado', index=False)

# Cargar el archivo Excel con openpyxl
wb = load_workbook(new_file_path)
ws = wb["Estado"]

# Definir el color rojo
data_fill = PatternFill(start_color='FE9072', end_color='FE9072', fill_type='solid')
mante_fill = PatternFill(start_color='F8CA78', end_color='F8CA78', fill_type='solid')
tigo_fill = PatternFill(start_color='71FF74', end_color='71FF74', fill_type='solid')
idaan_fill = PatternFill(start_color='89E7E7', end_color='89E7E7', fill_type='solid')
ensa_fill = PatternFill(start_color='EAE888', end_color='EAE888', fill_type='solid')
visa_fill = PatternFill(start_color='B8B8B8', end_color='B8B8B8', fill_type='solid')
seguro_fill = PatternFill(start_color='C41212', end_color='C41212', fill_type='solid')
netflix_fill = PatternFill(start_color='8992E7', end_color='8992E7', fill_type='solid')
anelys_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
naranja_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

# Obtener el índice de la columna "Descripcion"
desc_col_index = 0
for idx, cell in enumerate(ws[1]):
    if cell.value == 'Descripción':
        desc_col_index = idx + 1
        break
        
# Aplicar el color rojo a las filas que contengan la Descripcion "BANCA MOVIL ITBMS RECARGA TELEFONIA"
for row in ws.iter_rows(min_row=2):
    if row[desc_col_index - 1].value is not None:
        if "BANCA MOVIL ITBMS RECARGA TELEFONIA" in row[desc_col_index - 1].value:
            for cell in row:
                cell.fill = data_fill
        elif "BANCA MOVIL RECARGA TELEFONIA +MÓVIL" in row[desc_col_index - 1].value:
            for cell in row:
                cell.fill = data_fill
        elif "BANCA MOVIL TRANSFERENCIA A 0301011162515 ASAMBLEA DE PROPIETARIOS DEL P.H. ALTATERRA" in row[desc_col_index - 1].value:
            for cell in row:
                cell.fill = mante_fill
        elif "94827659" in row[desc_col_index - 1].value: #Tigo
            for cell in row:
                    cell.fill = tigo_fill
        elif "945168" in row[desc_col_index - 1].value: #IDAAN
            for cell in row:
                    cell.fill = tigo_fill
        elif "000021426294" in row[desc_col_index - 1].value: #ENSA
            for cell in row:
                    cell.fill = tigo_fill
        elif "356099660000" in row[desc_col_index - 1].value: #Data
            for cell in row:
                    cell.fill = tigo_fill
        elif "BANCA MOVIL PAGO VISA 4931-23XX-XXXX-9827 JEAN CARLOS COSSIO YEE" in row[desc_col_index - 1].value:
            for cell in row:
                    cell.fill = visa_fill
        elif "000000000300442612" in row[desc_col_index - 1].value:
            for cell in row:
                    cell.fill = seguro_fill
        elif "BANCA EN LINEA TRANSFERENCIA A 0472997148051 LUIS ALBERTO GONZALEZ REAL nefli" in row[desc_col_index - 1].value:
            for cell in row:
                    cell.fill = netflix_fill
        elif "METRO Y METROBUS " in row[desc_col_index - 1].value:
            for cell in row:
                    cell.fill = naranja_fill
        elif "BANCA EN LINEA TRANSFERENCIA DE ANELYS KRISCHELL GONZALEZ CHAVEZ" in row[desc_col_index - 1].value:
            for cell in row:
                    cell.fill = anelys_fill


# for row in ws.iter_rows(min_row=2):
#     if row[desc_col_index - 1].value is not None:
#         if "PAGO YAPPY A Pedidos Ya" in row[desc_col_index - 1].value:
#             for cell in row:
#                 cell.fill = data_fill
#         elif "BANCA MOVIL RECARGA TELEFONIA +MÓVIL" in row[desc_col_index - 1].value:
#             for cell in row:
#                 cell.fill = data_fill
#         elif "MANADA SIBARITA" in row[desc_col_index - 1].value:
#             for cell in row:
#                 cell.fill = mante_fill
#         elif "BANCA MÓVIL TIGO HOGAR - PYME (94827659)" in row[desc_col_index - 1].value:
#             for cell in row:
#                     cell.fill = tigo_fill
#         elif "GOOGLE  WyzeAppAndroid" in row[desc_col_index - 1].value:
#             for cell in row:
#                     cell.fill = idaan_fill
#         elif "TEXACO" in row[desc_col_index - 1].value:
#             for cell in row:
#                     cell.fill = ensa_fill
#         elif "SMARTFIT" in row[desc_col_index - 1].value:
#             for cell in row:
#                     cell.fill = visa_fill
#         elif "Amazon Prime" in row[desc_col_index - 1].value:
#             for cell in row:
#                     cell.fill = seguro_fill
#         elif "ACH - BGENERAL PLANILL" in row[desc_col_index - 1].value:
#             for cell in row:
#                     cell.fill = netflix_fill

# Guardar los cambios en el archivo Excel
wb.save(new_file_path)

